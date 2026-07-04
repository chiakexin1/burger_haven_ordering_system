import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, PieChart

def generate_sales_report(data_file):
    try:
        # Load order data
        orders_df = pd.read_excel(data_file, sheet_name="Customer_Order")
        if orders_df.empty:
            print("No sales data available.")
            return

        # Prepare date and month
        orders_df['Date'] = pd.to_datetime(orders_df['Date'], errors='coerce')
        orders_df.dropna(subset=['Date'], inplace=True)
        orders_df['Month'] = orders_df['Date'].dt.to_period('M')

        # === Summary ===
        print(f"Total Sales Amount: RM{orders_df['Order Amount'].sum():,.2f}")
        print(f"Total Number of Orders: {len(orders_df)}")

        # === Payment Method Summary ===
        payment_summary = orders_df['Payment Method'].value_counts().reset_index()
        payment_summary.columns = ["Payment Method", "Count"]
        print("\nPayment Method Breakdown:")
        print(payment_summary.to_string(index=False))

        # === Sales by Month ===
        month_sales = orders_df.groupby('Month')['Order Amount'].sum().reset_index()
        month_sales.columns = ["Month", "Sales Amount"]
        print("\nSales by Month:")
        print(month_sales.to_string(index=False, formatters={"Sales Amount": "RM{:,.2f}".format}))

        # === Monthly Customer Order Report ===
        customer_monthly = orders_df.groupby(['Customer Name', 'Month']).agg(
            Total_Spent=('Order Amount', 'sum'),
            Total_Orders=('Order Amount', 'count'),
            Products_Purchased=('Ordered Products', lambda x: ', '.join(x))
        ).reset_index()
        customer_monthly['Total_Spent'] = customer_monthly['Total_Spent'].apply(lambda x: f"RM{x:,.2f}")
        print("\nMonthly Customer Order Report:")
        print(customer_monthly.to_string(index=False))

        # === Top-Selling Products ===
        product_counts = {}
        for order in orders_df['Ordered Products']:
            for item in order.split(", "):
                name, qty = item.rsplit(' x', 1)
                product_counts[name] = product_counts.get(name, 0) + int(qty)

        top_products = pd.DataFrame(product_counts.items(), columns=["Product", "Quantity Sold"]).sort_values(by="Quantity Sold", ascending=False)
        print("\nTop-Selling Products:")
        print(top_products.to_string(index=False))

        # === Save reports to Excel ===
        with pd.ExcelWriter(data_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            top_products.to_excel(writer, sheet_name="Top_Selling_Products", index=False)
            month_sales.to_excel(writer, sheet_name="Sales_By_Month", index=False)
            payment_summary.to_excel(writer, sheet_name="Payment_Method_Breakdown", index=False)
            customer_monthly.to_excel(writer, sheet_name="Monthly_Customer_Report", index=False)

        # === Add charts ===
        wb = load_workbook(data_file)

        # Bar Chart: Top Products
        ws_products = wb["Top_Selling_Products"]
        chart1 = BarChart()
        chart1.title = "Top Selling Products"
        chart1.x_axis.title = "Product"
        chart1.y_axis.title = "Quantity Sold"
        data = Reference(ws_products, min_col=2, min_row=1, max_row=ws_products.max_row)
        cats = Reference(ws_products, min_col=1, min_row=2, max_row=ws_products.max_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        ws_products.add_chart(chart1, "E2")

        # Pie Chart: Payment Method Distribution
        ws_payment = wb["Payment_Method_Breakdown"]
        chart2 = PieChart()
        chart2.title = "Payment Method Distribution"
        data = Reference(ws_payment, min_col=2, min_row=1, max_row=ws_payment.max_row)
        labels = Reference(ws_payment, min_col=1, min_row=2, max_row=ws_payment.max_row)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(labels)
        ws_payment.add_chart(chart2, "D2")

        # Save workbook
        wb.save(data_file)
        print("\nSales report with charts generated and saved to Excel.")

    except Exception as e:
        print("Error generating sales report:", e)
