import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

def food_ordering_system(ws, wb, data_file):
    try:
       
        menu_df = pd.read_excel(data_file, sheet_name="MasterList")
        
        if menu_df.empty:
            print("Menu is currently empty. Please add burger products first.")
            return

        print("\nAvailable Menu:\n", menu_df[['Product ID', 'Product Name', 'Price']].to_string(index=False))

        # Customer Details with Validation
        cust_name = input("\nEnter Customer Name: ").strip()
        while not cust_name:
            cust_name = input("Customer Name cannot be empty. Please enter again: ").strip()

        while True:
            cust_id = input("Enter Customer ID (numeric): ").strip()
            if cust_id.isdigit():
                cust_id = int(cust_id)
                break
            print("Customer ID must be numeric. Please enter again.")

        address = input("Enter Delivery Address: ").strip()
        while not address:
            address = input("Delivery Address cannot be empty. Please enter again: ").strip()

        payment = input("Enter Payment Method (e.g., Cash, Card): ").strip()
        while not payment:
            payment = input("Payment Method cannot be empty. Please enter again: ").strip()

        order_date = datetime.now().strftime("%Y-%m-%d")
        order_time = datetime.now().strftime("%H:%M:%S")

        ordered_products_list = []
        total_amount = 0.0

        while True:
            pid = input("Enter Product ID (e.g., B001): ").strip().upper()
            if pid in menu_df['Product ID'].astype(str).values:
                try:
                    qty = int(input("Enter Quantity: "))
                    if qty <= 0:
                        print("Quantity must be greater than zero.")
                        continue
                except ValueError:
                    print("Invalid quantity. Please enter a valid number.")
                    continue

                item = menu_df[menu_df['Product ID'].astype(str) == pid].iloc[0]
                ordered_products_list.append(f"{item['Product Name']} x{qty}")
                total_amount += item['Price'] * qty
            else:
                print("Invalid Product ID.")

            more = input("Add more items? (Y/N): ").strip().upper()
            if more != 'Y':
                break

        if not ordered_products_list:
            print("No items were added to the order. Cancelling order.")
            return

        # Create a new order record
        new_order = {
            'Date': order_date,
            'Time': order_time,
            'Customer Name': cust_name,
            'Customer ID': cust_id,
            'Delivery Address': address,
            'Ordered Products': ", ".join(ordered_products_list),
            'Order Amount': round(total_amount, 2),
            'Payment Method': payment,
            'Order Delivery Status': "Pending",
            'Remarks': ""
        }

        order_df = pd.DataFrame([new_order])

        try:
            existing_orders = pd.read_excel(data_file, sheet_name="Customer_Order")
            existing_orders = existing_orders.dropna(axis=1, how='all')
            order_df = order_df.dropna(axis=1, how='all')
            order_df = pd.concat([existing_orders, order_df], ignore_index=True)
        except:
            pass  

        # Save the updated order list to the Excel file
        with pd.ExcelWriter(data_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            wb = load_workbook(data_file)
            if "Customer_Order" in wb.sheetnames:
                wb.remove(wb["Customer_Order"])
            wb.save(data_file)
            order_df.to_excel(writer, sheet_name="Customer_Order", index=False)

        print("\nOrder placed and saved successfully.\n")

    except Exception as e:
        print("Error:", e)
